"""
Second-pass placement: resolve duplicate titles by runtime.

    python scripts/match_by_duration.py                 report only
    python scripts/match_by_duration.py --write         fill in the folders

The first pass matched Bunny videos to files on disk by title, and refused
to guess whenever a title appeared more than once. That was the right call —
four publishers each ship a "Clostridium botulinum atf" — but it left 1,341
videos with no folder, so folders that hold seven lessons display one.

The Video Comparison sheet closes the gap: it carries a duration for every
same-title variant. Bunny reports a duration too, so within one title group
the copies can be paired by runtime instead of guessed.

Rules, deliberately conservative:

  * a pairing is accepted only if the best candidate is within TOLERANCE
    seconds and the runner-up is at least MARGIN seconds further away
  * each Bunny video is claimed at most once per group
  * only items that currently have no folder are touched, so nothing the
    first pass got right can be disturbed

Anything ambiguous stays unplaced and is counted in the report. A wrong
folder is worse than an absent one: it files a Pixorize video inside
Sketchy, where you would never think to question it.
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = Path(
    r"C:\Users\micha\OneDrive\Resource Guide AI\outputs"
    r"\019ff41f-0454-7e81-82f8-44dbf17b8fca\Medical_Resource_Library_Index.xlsx"
)
CATALOG = ROOT / "data" / "catalog.json"
BUNNY = ROOT / "bunny_catalog.json"

WRITE = "--write" in sys.argv

# A transcode shifts runtime slightly; anything beyond this is a different video.
TOLERANCE = 4.0
# The runner-up must be at least this much further away, or the pair is a coin flip.
MARGIN = 3.0

BREADCRUMB_SEP = re.compile(r"\s*[\u203a>/]\s*")


def norm(title: str) -> str:
    title = re.sub(r"\.(mp4|m4v|mkv|mov|webm|avi|wmv|flv|ts|mpg|mpeg)$", "", str(title or ""), flags=re.I)
    return re.sub(r"[^a-z0-9]+", " ", title.lower()).strip()


def read_comparison_sheet() -> list[dict]:
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb[next(s for s in wb.sheetnames if "comparison" in s.lower())]

    rows = ws.iter_rows(values_only=True)
    header = None
    out = []
    for row in rows:
        if header is None:
            if row and row[0] == "Group ID":
                header = {name: i for i, name in enumerate(row) if name}
            continue
        if not row or not row[0]:
            continue
        get = lambda key: row[header[key]] if key in header else None  # noqa: E731
        crumb = str(get("Folder Breadcrumb") or "")
        parts = [p.strip() for p in BREADCRUMB_SEP.split(crumb) if p.strip()]
        try:
            seconds = float(get("Duration (sec)"))
        except (TypeError, ValueError):
            continue
        out.append({
            "group": get("Group ID"),
            "filename": str(get("Exact Filename") or ""),
            "folder": "/".join(parts),
            "seconds": seconds,
        })
    return out


def main() -> int:
    if not XLSX.exists():
        print(f"Workbook not found: {XLSX}")
        return 1

    disk = read_comparison_sheet()
    print(f"Same-title variants on disk : {len(disk):,}")

    bunny_raw = json.loads(BUNNY.read_text(encoding="utf-8"))
    catalog = json.loads(CATALOG.read_text(encoding="utf-8"))
    by_id = {i["id"]: i for i in catalog["items"]}

    # Bunny videos that are in the catalogue and still have no folder.
    unplaced: dict[str, list[dict]] = {}
    for v in bunny_raw:
        item = by_id.get(v["id"])
        if not item or item.get("folder"):
            continue
        secs = v.get("duration_seconds")
        if not secs:
            continue
        unplaced.setdefault(norm(v["title"]), []).append({"id": v["id"], "seconds": float(secs)})

    # Disk copies grouped the same way.
    groups: dict[str, list[dict]] = {}
    for d in disk:
        if d["folder"]:
            groups.setdefault(norm(d["filename"]), []).append(d)

    resolved: dict[str, str] = {}       # bunny id -> folder
    stats = {"groups": 0, "paired": 0, "declined_margin": 0, "declined_tolerance": 0,
             "no_candidates": 0, "tie_same_collection": 0}

    for key, copies in groups.items():
        candidates = unplaced.get(key)
        if not candidates:
            stats["no_candidates"] += len(copies)
            continue
        stats["groups"] += 1

        # When every copy of a title sits inside one collection, the copies are
        # the same lesson filed twice by the same publisher — Sketchy keeps one
        # under Microbiology and another under Old but Gold. Identical runtimes
        # make them undecidable, but also interchangeable: either folder is
        # correct, so a tie can be settled arbitrarily. Across publishers it
        # stays undecidable and stays unplaced.
        one_collection = len({c["folder"].split("/")[0] for c in copies}) == 1

        taken: set[str] = set()
        # Longest first: the biggest runtime gaps are the least ambiguous, so
        # settling them early stops a confident pair losing its video to a
        # marginal one.
        for d in sorted(copies, key=lambda c: (-c["seconds"], c["folder"])):
            ranked = sorted(
                (c for c in candidates if c["id"] not in taken),
                key=lambda c: abs(c["seconds"] - d["seconds"]),
            )
            if not ranked:
                stats["no_candidates"] += 1
                continue

            best = ranked[0]
            delta = abs(best["seconds"] - d["seconds"])
            if delta > TOLERANCE:
                stats["declined_tolerance"] += 1
                continue
            if len(ranked) > 1 and not one_collection:
                runner_up = abs(ranked[1]["seconds"] - d["seconds"])
                if runner_up - delta < MARGIN:
                    stats["declined_margin"] += 1
                    continue
            if one_collection and len(ranked) > 1:
                runner_up = abs(ranked[1]["seconds"] - d["seconds"])
                if runner_up - delta < MARGIN:
                    stats["tie_same_collection"] += 1

            resolved[best["id"]] = d["folder"]
            taken.add(best["id"])
            stats["paired"] += 1

    print(f"Title groups considered     : {stats['groups']:,}")
    print(f"  paired by runtime         : {stats['paired']:,}")
    print(f"  declined, too far apart   : {stats['declined_tolerance']:,}")
    print(f"  declined, too close to call: {stats['declined_margin']:,}  (different publishers)")
    print(f"  tie settled within one collection: {stats['tie_same_collection']:,}")
    print(f"  no unplaced Bunny copy    : {stats['no_candidates']:,}")

    if not WRITE:
        print("\nReport only. Re-run with --write to apply.")
        return 0

    slug = lambda s: re.sub(r"^-+|-+$", "", re.sub(r"[^a-z0-9]+", "-", s.lower()))  # noqa: E731

    import shutil
    shutil.copy(CATALOG, str(CATALOG) + ".bak")

    applied = 0
    for vid, folder in resolved.items():
        item = by_id.get(vid)
        if not item or item.get("folder"):
            continue
        parts = folder.split("/")
        item["folder"] = folder
        item["collection"] = parts[0]
        item["section"] = parts[1] if len(parts) > 1 else ""
        item["tags"] = sorted(set(item.get("tags", [])) | {slug(p) for p in parts})
        applied += 1

    # Rebuild the folder tree and collection facets from the enriched items.
    root: dict = {}
    for item in catalog["items"]:
        if not item.get("folder"):
            continue
        level = root
        trail = []
        for name in item["folder"].split("/"):
            trail.append(name)
            node = level.setdefault(name, {"label": name, "path": "/".join(trail), "count": 0, "children": {}})
            node["count"] += 1
            level = node["children"]

    def to_tree(level):
        return sorted(
            ({"label": n["label"], "path": n["path"], "count": n["count"],
              "children": to_tree(n["children"])} for n in level.values()),
            key=lambda n: -n["count"],
        )

    catalog["folders"] = to_tree(root)

    tag_counts: dict[str, int] = {}
    for item in catalog["items"]:
        for t in item.get("tags", []):
            tag_counts[t] = tag_counts.get(t, 0) + 1
    labels = {f["id"]: f["label"] for f in catalog.get("tagFacets", [])}
    catalog["tagFacets"] = sorted(
        ({"id": k, "label": labels.get(k, k), "count": v} for k, v in tag_counts.items()),
        key=lambda t: (-t["count"], t["id"]),
    )
    catalog["tags"] = [t["id"] for t in catalog["tagFacets"]]

    CATALOG.write_text(json.dumps(catalog), encoding="utf-8")
    print(f"\nApplied folders to {applied:,} videos. Backup at catalog.json.bak")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
